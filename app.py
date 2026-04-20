import openpyxl
import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
from datetime import datetime
import re
from openpyxl.styles import PatternFill, Border, Side

######################### Constant variables #########################

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

valid_reports = ["ח", "ג", "נ", "י", "0", "1", "ימי התארגנות", "משתחרר"]
valid_locations = ["אביבים", "תורמוס", "יבא תחתון", "רייצ\'ל", "דובב", "קיסוסית", "מארון א ראס", "שקד", "ירון"]

valid_values = valid_reports + valid_locations

valid_roles = ["קצין", "מט\"ב", "חובש", "נהג משא כבד", "גשש", "נהג קו", "נהג טנק", "טען", "תותחן", "מט\"ק", "לוחם",
               "יתר", "מפקד D9", "מפעיל D9"]
valid_sosh = ["סדיר", "מילואים"]

date_columns = ["תאריך התייצבות", "תאריך סיום", "ת.סיום התארגנות"]

valid_status = ["V", "שוחרר", "התארגנות", "סיים שמ\"פ"]
shamap_status = ["V", "התארגנות"]
shamap_reports = valid_locations + ["ח", "ג", "י", "1", "ימי התארגנות", "משתחרר"]

DEPARTMENTS_LIST = ["פלוגה מבצעית א'", "פלוגה מבצעית ב'", "פלוגה מבצעית ג'", "פלוגה מסייעת", "מפג\"ד", "פלס\"ם",
                    "אג\"מ"]

# עמודות שנרצה להשוות בין השבצק הגדודי לפלוגתי
fields_to_check = ["שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "התייצב", "סו\"ש"]  # תפקיד נבדק בנפרד

# valid_AGAM = ["ח", "ג", "נ", "י", "0", "עתניאל" ,"1", "ימי התארגנות"]
# valid_PALSAM = ["ח", "ג", "נ", "י", "0", "עתניאל", "אדוריים", "630", "620", "710", "720", "חורסה", "כרם נגוהות", "בית חגי", "סוסיא", "חוות מעון", "1", "ימי התארגנות"]
# valid_MAFGAD = ["ח", "ג", "נ", "י", "0", "עתניאל", "1", "ימי התארגנות"]
# valid_A = ["ח", "ג", "נ", "י", "0", "אדוריים", "חורסה", "כרם נגוהות", "1", "ימי התארגנות"]
# valid_B = ["ח", "ג", "נ", "י", "0", "סוסיא", "720", "1", "חוות מעון", "ימי התארגנות"]
# valid_C = ["ח", "ג", "נ", "י", "0", "710", "620","בית חגי", "1", "ימי התארגנות"]
# valid_D = ["ח", "ג", "נ", "י", "0", "עתניאל", "630", "1", "ימי התארגנות"]

######################### Mapping functions #########################

FILENAME_TO_DEPARTMENT = {
    "שבצק פלוגה א'": "פלוגה מבצעית א'",
    "שבצק פלוגה ב'": "פלוגה מבצעית ב'",
    "שבצק פלוגה ג'": "פלוגה מבצעית ג'",
    "שבצק מסייעת": "פלוגה מסייעת",
    "שבצק אג\"מ": "אג\"מ",
    "שבצק אג_מ": "אג\"מ",
    "שבצק פלס\"ם": "פלס\"ם",
    "שבצק פלס_ם": "פלס\"ם",
    "שבצק מפג\"ד": "מפג\"ד",
    "שבצק מפג_ד": "מפג\"ד",
}

requests_TO_DEPARTMENT = {
    "א'": "פלוגה מבצעית א'",
    "ב'": "פלוגה מבצעית ב'",
    "ג'": "פלוגה מבצעית ג'",
    "מסייעת": "פלוגה מסייעת",
    "אג\"ם": "אג\"מ",
    "פלס\"ם": "פלס\"ם",
    "מפג\"ד": "מפג\"ד"
}

# UNIT_VALID_MAP = {
#     "אג\"מ": valid_AGAM,
#     "אג_מ": valid_AGAM,
#     "אגמ": valid_AGAM,
#     "פלסם": valid_PALSAM,
#     "פלס\"ם": valid_PALSAM,
#     "פלס_ם": valid_PALSAM,
#     "מפגד": valid_MAFGAD,
#     "מפג_ד": valid_MAFGAD,
#     "מפג\"ד": valid_MAFGAD,
#     "פלוגה א'": valid_A,
#     "פלוגה ב'": valid_B,
#     "פלוגה ג'": valid_C,
#     "פלוגה מסייעת": valid_D,
# }

# dep_map = {
#     "מפקדת היחידה": "מפג\"ד",
#     "פלוגה מבצעית א'": "פלוגה א'",
#     "פלוגה מבצעית ב'": "פלוגה ב'",
#     "פלוגה מבצעית ג'": "פלוגה ג'",
#     "פלוגה מבצעית ד'": "פלוגה מסייעת'"
# }

COMMENT_PRIORITY = {
    "High": "#f8d7da",  # pastel red
    "Medium": "#fff3cd",  # pastel yellow
    "Low": None
}


############################# Functions #####################################

################# General #################

def go_to(page_name):
    st.session_state.page = page_name


def format_cell(val):
    if pd.isna(val):
        return ""
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%d/%m/%y")
    return str(val).strip() if not isinstance(val, float) else str(int(val)).strip()


def format_report_column_name(col):
    col_date = parse_column_date(col)

    if pd.notna(col_date):
        return col_date.strftime("%d/%m/%y")

    return str(col).strip()


# def get_valid_values_by_filename(filename: str):
#     filename = filename.lower()
#     for key, valid_list in UNIT_VALID_MAP.items():
#         if key.replace('"', '').lower() in filename:
#             return valid_list
#     return valid_values
def render_comments_table(df):
    if df is None or df.empty:
        st.info("אין הערות")
        return

    styled_rows = []
    for _, row in df.iterrows():
        color = row.get("Color")
        bg = f'background-color:{color};' if color else ''
        styled_rows.append(
            f"<tr style='{bg}'>" +
            "".join(f"<td style='white-space: nowrap; padding: 6px 10px;'>{row[col]}</td>"
                    for col in ["מ.א", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "הערה"]) +
            "</tr>"
        )

    html = f"""
    <div style='max-height: 600px; overflow-y: auto; overflow-x: auto; border: 1px solid #ccc; border-radius: 6px;'>
        <table style='border-collapse: collapse; width: 100%; direction: rtl; font-size: 14px;'>
            <thead>
                <tr style='background-color: #f0f0f0; text-align: right;'>
                    <th style='padding: 8px;'>מ.א</th>
                    <th style='padding: 8px;'>שם פרטי</th>
                    <th style='padding: 8px;'>שם משפחה</th>
                    <th style='padding: 8px;'>מסגרת ראשית</th>
                    <th style='padding: 8px;'>מסגרת משנית</th>
                    <th style='padding: 8px;'>הערה</th>
                </tr>
            </thead>
            <tbody>
                {''.join(styled_rows)}
            </tbody>
        </table>
    </div>
    """

    st.markdown(html, unsafe_allow_html=True)


################# Daily Update #################
BASE_TEXT_COLUMNS = [
    "מ.א", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית",
    "התייצב", 'סו"ש', 'סו\"ש', "תפקיד", "מסופח"
]


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # ניקוי שמות עמודות
    df.columns = [str(c).strip() for c in df.columns]

    # איחוד שם עמודה סו"ש
    if 'סו\"ש' in df.columns and 'סו"ש' not in df.columns:
        df = df.rename(columns={'סו\"ש': 'סו"ש'})

    # עמודות טקסט
    for col in BASE_TEXT_COLUMNS:
        if col in df.columns:
            df[col] = df[col].astype("object")

    # עמודות תאריך קבועות
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

    # עמודות דיווח יומיות = כל כותרת שהיא תאריך
    for col in df.columns:
        if pd.notna(parse_column_date(col)):
            df[col] = df[col].astype("object")

    return df


def safe_set_cell(df: pd.DataFrame, idx, col, value):
    if col not in df.columns:
        raise KeyError(f"Column '{col}' does not exist")

    if pd.isna(value):
        df.at[idx, col] = pd.NA
        return

    if pd.notna(parse_column_date(col)):
        df.at[idx, col] = normalize_report_value_for_excel(value)
        return

    if col in BASE_TEXT_COLUMNS:
        df.at[idx, col] = str(value).strip()
        return

    df.at[idx, col] = value


def is_valid_id(id_number):
    if pd.isna(id_number):
        return False
    id_str = str(id_number).strip()
    return id_str.isdigit() and len(id_str) == 7


# ---- Compare two shabzaks ---
def get_soldier_info(row):
    return [
        str(row.get("שם פרטי", "")).strip(),
        str(row.get("שם משפחה", "")).strip(),
        str(row.get("מסגרת ראשית", "")).strip(),
        str(row.get("מסגרת משנית", "")).strip()
    ]


def is_valid_value_for_column(col: str, val: str) -> bool:
    val = str(val).strip()
    if col == "תפקיד":
        return val in valid_roles
    elif col in ['סו"ש', 'סו\"ש']:
        return val in valid_sosh
    elif col == "התייצב":
        return val in valid_status
    return True


def is_row_empty(row):
    key_fields = ["מ.א", "שם פרטי", "שם משפחה"]

    for field in key_fields:
        val = row.get(field, None)

        if pd.isna(val):
            continue

        if isinstance(val, str):
            normalized = val.strip().lower()
            if normalized in ["", "nan", "none"]:
                continue

        return False

    return True


# def normalize_date(value):
#     if pd.isna(value):
#         return None
#     if isinstance(value, (datetime, pd.Timestamp)):
#         return value.strftime("%d/%m/%y")  # כבר אובייקט תאריך
#     try:
#         return datetime.strptime(str(value).strip(), "%d/%m/%y").strftime("%d/%m/%y")
#     except Exception:
#         return None
def handle_missing_in_central(df_central, dep_row):
    id_number = str(dep_row["מ.א"]).strip()
    central_match = df_central[df_central["מ.א"] == id_number]

    if central_match.empty:
        if not (id_number.isdigit() and len(id_number) == 7):
            return df_central, [id_number] + get_soldier_info(dep_row) + ["מ.א לא תקין, לא התבצע עדכון"], None

        common_cols = [col for col in df_central.columns if col in dep_row and col != "מ.א"]
        new_row = {col: dep_row[col] for col in common_cols}
        new_row["מ.א"] = id_number
        df_central = pd.concat([df_central, pd.DataFrame([new_row])], ignore_index=True)

        return df_central, [id_number] + get_soldier_info(dep_row) + [
            "לא נמצא בגדודי – נוסף על בסיס הפלוגתי: לוודא סטטוס ותאריך התייצבות"], None

    return df_central, None, central_match.index[0]


def analyze_suspicious_cases(df_central, df_dep, central_idx, dep_idx, col, central_val, dep_val, run_date):
    suspicious_comments = []
    dep_status = format_cell(df_dep.at[dep_idx, "התייצב"])
    central_status = format_cell(df_central.at[central_idx, "התייצב"])
    m_col = format_report_column_name(col)
    col_date = pd.to_datetime(col, errors="coerce")
    run_date = pd.Timestamp(run_date).normalize()
    if pd.notna(col_date) and col_date.normalize() == run_date:
        if dep_val == "ג" or central_val == "ג":
            if dep_status not in shamap_status or central_status not in shamap_status:
                suspicious_comments.append(
                    f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', לא דווח שהתייצב – לבדוק גימלים")
            else:
                suspicious_comments.append(
                    f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', דווח גימלים – לבדוק אישור רופא")

        if dep_val == "נ" or central_val == "נ":
            suspicious_comments.append(f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', שים לב לנפקדות")

    # if dep_val in ["0", "ימי התארגנות"] or central_val in ["0", "ימי התארגנות"]:
    #     if dep_status == "V" or central_status == "V":
    #         col_date = pd.to_datetime(col, errors="coerce")
    #         today = pd.Timestamp.today().normalize()
    #
    #         if pd.notna(col_date) and col_date.normalize() == today:
    #
    #
    #         # אם התאריך של 0/ימי התארגנות הוא אחרי או שווה לתאריך התייצבות → נחשב סיום שמפ
    #         if m_col_date >= enlist_date:
    #             df_central.at[central_idx, "התייצב"] = "התייצב ושוחרר"
    #             df_dep.at[dep_idx, "התייצב"] = "התייצב ושוחרר"
    #             suspicious_comments.append(
    #                 f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', סיום שמפ אתמול – ודא הזדכות על ציוד"
    #             )

    return suspicious_comments


def get_yesterday_value(df, row_idx, run_date):
    run_date = pd.Timestamp(run_date).normalize()
    yesterday_date = run_date - pd.Timedelta(days=1)

    # חיפוש עמודה שתואמת בדיוק לאתמול
    for col in df.columns:
        col_date = parse_column_date(col)

        if pd.notna(col_date) and col_date.normalize() == yesterday_date:
            val = format_cell(df.at[row_idx, col])

            if val in ["", "nan", "None", "NaT"]:
                return "0"

            return val

    # אם אין עמודה של אתמול
    return "0"


def detect_new_finish_shamap(df_central, today_val, central_idx, col, central_val, dep_val, run_date):
    comments = []

    col_date = parse_column_date(col)
    run_date = pd.Timestamp(run_date).normalize()

    if pd.isna(col_date) or col_date.normalize() != run_date:
        return comments

    m_col = format_report_column_name(col)

    yesterday_val = get_yesterday_value(df_central, central_idx, run_date)

    # --- התייצבות חדשה ---
    if yesterday_val not in shamap_reports and today_val in shamap_reports:

        if today_val == "ימי התארגנות":
            comments.append(
                f"{m_col}: אתמול='{yesterday_val}', היום='{today_val}' :ימי התארגנות ולא היה בשמפ אתמול - לברר עם הפלוגה")
            safe_set_cell(df_central, central_idx, "התייצב", "התארגנות")
        elif today_val == "ג":
            comments.append(
                f"{m_col}: אתמול='{yesterday_val}', היום='{today_val}' :דיווח גימלים לא בשמ\"פ - לברר עם הפלוגה")
        elif today_val == "משתחרר":
            comments.append(
                f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}'  מסיים שמ\"פ- לוודא שאכן התייצב לחד יומי ולעדכן ת.סיום")
        else:
            comments.append(
                f"{m_col}: אתמול='{yesterday_val}', היום='{today_val}' :התייצבות חדשה- לעדכן ת.התייצבות")
            safe_set_cell(df_central, central_idx, "התייצב", "V")

    # --- אתמול משתחרר והיום שוב בשמ"פ ---
    if yesterday_val == "משתחרר" and today_val not in ["ימי התארגנות", "0", ""]:
        comments.append(
            f"{m_col}: אתמול שוחרר והיום בשמ\"פ - לוודא שחרור מול הפלוגה"
        )

    # --- סיום שמ"פ ---
    if yesterday_val in shamap_reports and today_val not in shamap_reports:
        comments.append(
            f"{m_col}: אתמול='{yesterday_val}', היום='{today_val}' :נדרש לסגור שמ\"פ - עודכן סטטוס התייצבות לסיום שמפ. לעדכן את הקישור, לעדכן ת.סיום ו/או ת.סיום התארגנות "
        )
        safe_set_cell(df_central, central_idx, "התייצב", "סיים שמ\"פ")

    return comments


def compare_and_update_cell(df_central, df_dep, central_idx, dep_idx, col, run_date):
    comments = []
    central_raw = df_central.at[central_idx, col]
    dep_raw = df_dep.at[dep_idx, col]

    central_val = format_cell(central_raw)
    dep_val = format_cell(dep_raw)

    status_central = format_cell(df_central.at[central_idx, "התייצב"])
    status_dep = format_cell(df_dep.at[dep_idx, "התייצב"])

    # on_shamap = status_central in shamap_status or status_dep in shamap_status
    on_shamap = status_central in shamap_status

    is_central_empty = central_val in ["", "nan"]
    is_dep_empty = dep_val in ["", "nan"]

    # --- בדיקות חשודות ---
    comments.extend(
        analyze_suspicious_cases(df_central, df_dep, central_idx, dep_idx, col, central_val, dep_val, run_date))

    m_col = format_report_column_name(col)

    # --- גדודי ריק פלוגתי מלא ---
    if is_central_empty and not is_dep_empty:
        if dep_val == "ימי התארגנות" and status_central != "התארגנות":
            comments.append(
                f"{m_col}: פלוגתי='{dep_val}', דיווח ימי התארגנות - סטטוס התייצבות עודכן להתארגנות, לוודא תאריכי התארגנות")
            safe_set_cell(df_central, central_idx, "התייצב", "התארגנות")

        if dep_val not in valid_values:
            comments.append(
                f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', ערך לא חוקי בפלוגתי – לא בוצע עדכון, לברר עם הפלוגה")
        else:
            coms = detect_new_finish_shamap(
                df_central,
                dep_val,
                central_idx,
                col,
                central_val,
                dep_val,
                run_date
            )
            if coms:
                comments.extend(coms)
            safe_set_cell(df_central, central_idx, col, dep_raw)



    # --- גדודי מלא ופלוגתי מלא, לא תואמים ---
    elif not is_central_empty and not is_dep_empty and not (central_val == dep_val):
        if central_val != "ימי התארגנות":
            if dep_val not in valid_values:
                comments.append(
                    f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', ערך לא חוקי בפלוגתי – לא בוצע עדכון, לברר עם הפלוגה")
            else:

                coms = detect_new_finish_shamap(
                    df_central,
                    dep_val,
                    central_idx,
                    col,
                    central_val,
                    dep_val,
                    run_date
                )
                if coms:
                    comments.extend(coms)
                safe_set_cell(df_central, central_idx, col, dep_raw)

    # --- גדודי מלא ופלוגתי ריק ---
    elif not is_central_empty and is_dep_empty:
        if central_val not in valid_values:
            comments.append(
                f"{m_col}: גדודי='{central_val}', פלוגתי='{dep_val}', ערך לא חוקי בגדודי ואין דיווח בפלוגתי- לברר עם הפלוגה")
        else:

            coms = detect_new_finish_shamap(
                df_central,
                central_val,
                central_idx,
                col,
                central_val,
                dep_val,
                run_date
            )

        if coms:
            comments.extend(coms)

    # --- גדודי ריק ופלוגתי ריק אבל החייל בשמ"פ ---
    elif is_central_empty and is_dep_empty and on_shamap:
        enlistment_date_str = df_central.at[central_idx, "תאריך התייצבות"]
        try:
            # פירוש תאריכים בפורמט יום/חודש/שנה
            enlistment_date = pd.to_datetime(enlistment_date_str, errors="coerce", dayfirst=True)
            col_date = pd.to_datetime(col, errors="coerce", dayfirst=True)
            runing_date = pd.Timestamp(run_date).normalize().date()
            if pd.notna(col_date) and pd.notna(enlistment_date):
                col_date = col_date.date()
                enlistment_date = enlistment_date.date()
                if enlistment_date <= col_date <= runing_date:
                    comments.append(f"{m_col}: דווח שהתייצב, אך חסר דיווח")
            else:
                comments.append(f"{m_col}: תאריך התייצבות לא תקין")
        except Exception as e:
            print("❌ ERROR in התייצבות check:", e)

    return comments


#########################################

def compare_shared_basic_fields(df_central, df_dep, common_ids):
    comments = []

    for col in fields_to_check:
        if col in df_central.columns:
            df_central[col] = df_central[col].astype("object")
        if col in df_dep.columns:
            df_dep[col] = df_dep[col].astype("object")

    if "תפקיד" in df_central.columns:
        df_central["תפקיד"] = df_central["תפקיד"].astype("object")

    central_idx_map = df_central.reset_index().set_index("מ.א")["index"].to_dict()
    dep_idx_map = df_dep.reset_index().set_index("מ.א")["index"].to_dict()

    for id_number in common_ids:
        if not is_valid_id(id_number):
            continue

        central_idx = central_idx_map.get(id_number)
        dep_idx = dep_idx_map.get(id_number)

        if central_idx is None or dep_idx is None:
            continue

        dep_info = get_soldier_info(df_dep.loc[dep_idx])

        for col in fields_to_check:
            if col not in df_central.columns or col not in df_dep.columns:
                continue

            central_val = format_cell(df_central.at[central_idx, col])
            dep_val = format_cell(df_dep.at[dep_idx, col])

            central_empty = central_val in ["", "nan", "None", "NaT"]
            dep_empty = dep_val in ["", "nan", "None", "NaT"]

            # גדודי ריק, פלוגתי מלא - נעדכן לפי פלוגתי רק אם ערך חוקי
            if central_empty and not dep_empty:
                if is_valid_value_for_column(col, dep_val):
                    df_central.at[central_idx, col] = dep_val
                    comments.append([id_number] + dep_info + [f"{col}: ריק בגדודי, הושלם לערך '{dep_val}'"])
                else:
                    comments.append([id_number] + dep_info + [
                        f"{col}: ריק בגדודי, ערך לא חוקי בפלוגתי ('{dep_val}') – לא הושלם לגדודי"
                    ])

            # פלוגתי ריק, גדודי מלא - רק בדיקת תקינות של הגדודי
            elif dep_empty and not central_empty:
                if not is_valid_value_for_column(col, central_val):
                    comments.append([id_number] + dep_info + [
                        f"{col}: פלוגתי ריק, ערך לא חוקי בגדודי ('{central_val}')"
                    ])

            # פלוגתי וגדודי מלאים אבל שונים - רק מסגרת משנית מתעדכנת לפי הפלוגתי
            elif not central_empty and not dep_empty and central_val != dep_val:
                if col == "מסגרת משנית" and is_valid_value_for_column(col, dep_val):
                    df_central.at[central_idx, col] = dep_val
                    comments.append(
                        [id_number] + dep_info + [f"{col}: חוסר התאמה – עודכן לפי הפלוגתי ('{dep_val}')"]
                    )

        # בדיקה נפרדת של תפקיד - רק בגדודי
        if "תפקיד" in df_central.columns:
            central_role = format_cell(df_central.at[central_idx, "תפקיד"])
            central_role_empty = central_role in ["", "nan", "None", "NaT"]

            if not central_role_empty and central_role not in valid_roles:
                comments.append(
                    [id_number] + dep_info + [f"תפקיד: ערך לא חוקי בגדודי ('{central_role}')"]
                )

    return comments


def get_department_from_filename(filename):
    filename = Path(filename).stem.strip()
    for key, dept in FILENAME_TO_DEPARTMENT.items():
        if key in filename:
            return dept
    return None  # לא מזוהה


def clean_id_column(series):
    return series.apply(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() else str(x).strip())


def check_duplicate_ids(df_central, df_dep):
    comments = []
    duplicated_central = df_central[df_central.duplicated("מ.א", keep=False)]
    duplicated_dep = df_dep[df_dep.duplicated("מ.א", keep=False)]

    for _, row in duplicated_central.iterrows():
        comments.append([row["מ.א"]] + get_soldier_info(row) + ["כפילות מ.א בקובץ גדודי – נדרש בירור"])

    for _, row in duplicated_dep.iterrows():
        comments.append([row["מ.א"]] + get_soldier_info(row) + ["כפילות מ.א בקובץ פלוגתי – נדרש בירור"])

    return comments


def check_valid_departments(df, source_name):
    comments = []

    for idx, row in df.iterrows():

        raw_val = str(row.get("מסגרת ראשית", "")).strip()

        if raw_val not in DEPARTMENTS_LIST:
            comments.append(
                [row.get("מ.א", "")] + get_soldier_info(row) + [f"'{raw_val}' אינה מסגרת חוקית ({source_name})"])

    return comments


def add_missing_from_dep(df_central, df_dep, ids_central, ids_dep, common_cols):
    comments = []
    missing_in_central = ids_dep - ids_central

    for id_number in missing_in_central:
        row = df_dep[df_dep["מ.א"] == id_number].iloc[0]

        if not is_valid_id(id_number):
            comments.append([id_number] + get_soldier_info(row) + ["מ.א לא תקין - לא נוסף לשבצק"])
            continue
        new_row = {col: row[col] for col in common_cols}
        new_row["מ.א"] = id_number
        df_central = pd.concat([df_central, pd.DataFrame([new_row])], ignore_index=True)
        comments.append([id_number] + get_soldier_info(row) + [
            "לא נמצא בגדודי – נוסף על בסיס הפלוגתי - לוודא סטטוס ותאריך התייצבות"])

    return df_central, comments


def add_missing_from_central(df_dep, df_central, ids_dep, ids_central, common_cols, dep_filename):
    comments = []
    missing_in_dep = ids_central - ids_dep

    current_dep_name = get_department_from_filename(dep_filename)

    for id_number in missing_in_dep:
        row = df_central[df_central["מ.א"] == id_number].iloc[0]
        if not is_valid_id(id_number):
            comments.append([id_number] + get_soldier_info(row) + [
                "מ.א לא תקין - לא נוסף לשבצק - לתקן ולהוסיף לפלוגתי דרך האפליקציה"])
            continue
        unit = str(row.get("מסגרת ראשית", "")).strip()
        status = str(row.get("התייצב", "")).strip()

        if current_dep_name and unit == current_dep_name and status in shamap_status:
            new_row = {col: row[col] for col in common_cols}
            new_row["מ.א"] = id_number
            df_dep = pd.concat([df_dep, pd.DataFrame([new_row])], ignore_index=True)
            comments.append([id_number] + get_soldier_info(row) + ["לא נמצא בפלוגתי - להוסיף דרך האפליקציה"])

    return df_dep, comments


################# Excel #################
def apply_excel_formatting(worksheet, df):
    worksheet.sheet_view.rightToLeft = True  # ✅ יישור מימין לשמאל

    # הגדרות צבעים וגבולות
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # כחול בהיר
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    date_format = "DD/MM/YYYY"

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                                      min_col=1, max_col=worksheet.max_column), start=1):
        for col_idx, cell in enumerate(row, start=1):
            header_name = worksheet.cell(row=1, column=col_idx).value
            parsed_header_date = pd.to_datetime(header_name, errors="coerce", dayfirst=True)

            # צבע רקע לשורת כותרת
            if row_idx == 1:
                cell.fill = header_fill
            else:
                cell.fill = PatternFill(fill_type=None)  # רקע לבן / שקוף

            # גבולות
            cell.border = border

            # פורמט תאריך לעמודות תאריך
            if row_idx > 1 and header_name in date_columns:
                cell.number_format = date_format
            if row_idx == 1 and pd.notna(parsed_header_date):
                cell.number_format = date_format


def merge_all_sheets(uploaded_file):
    try:
        # טען את כל הגיליונות כ־DataFrames
        all_sheets = pd.read_excel(uploaded_file, sheet_name=None)  # מחזיר dict: sheet_name -> DataFrame

        merged_df = pd.DataFrame()

        for sheet_name, df in all_sheets.items():
            df = df.copy()
            if not df.empty:
                df["sheet"] = sheet_name
                merged_df = pd.concat([merged_df, df], ignore_index=True)

        return merged_df

    except Exception as e:
        st.error(f"שגיאה באיחוד הגיליונות: {e}")
        return None


def split_to_sheets(df):
    df = df.copy()

    if "sheet" not in df.columns or "מסגרת משנית" not in df.columns:
        raise ValueError("העמודות 'sheet' ו-'מסגרת משנית' חייבות להופיע בטבלה")

    existing_sheets = set(df["sheet"].dropna().astype(str).str.strip().unique())

    for idx, row in df[df["sheet"].isna()].iterrows():
        alt_val = str(row["מסגרת משנית"]).strip()

        if alt_val in existing_sheets:
            df.at[idx, "sheet"] = alt_val
        else:
            df.at[idx, "sheet"] = "נוספים"

    result = {}
    for sheet_name, group in df.groupby("sheet"):
        result[sheet_name] = group.drop(columns=["sheet"])  # הסרה של העמודה מהתוצאה אם רוצים

    return result


def normalize_report_value_for_excel(val):
    if pd.isna(val):
        return ""

    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()

    if isinstance(val, int):
        return str(val)

    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%d/%m/%Y")

    val_str = str(val).strip()

    try:
        num = float(val_str)
        if num.is_integer():
            return str(int(num))
    except Exception:
        pass

    return val_str


def prepare_date_columns_for_excel(df):
    df = df.copy()

    for col in df.columns:
        parsed_col_date = parse_column_date(col)

        if pd.notna(parsed_col_date):
            df[col] = df[col].apply(normalize_report_value_for_excel)

        elif col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

    return df


def convert_excel_headers(df):
    df = df.copy()
    new_columns = []

    for col in df.columns:
        parsed_col_date = parse_column_date(col)
        if pd.notna(parsed_col_date):
            new_columns.append(parsed_col_date.to_pydatetime())
        else:
            new_columns.append(col)

    df.columns = new_columns
    return df


def to_excel_bytes(df, sheet_name):
    df = prepare_date_columns_for_excel(df)
    df = convert_excel_headers(df)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        apply_excel_formatting(worksheet, df)  # ✅ עיצוב מלא
    return output.getvalue()


def dict_to_excel_bytes(sheets_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df = prepare_date_columns_for_excel(df)
            safe_sheet_name = str(sheet_name)[:31]
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
            worksheet = writer.sheets[safe_sheet_name]
            apply_excel_formatting(worksheet, df)  # ✅ עיצוב מלא
    return output.getvalue()


def to_colored_excel(df, sheet_name):
    df = prepare_date_columns_for_excel(df)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.drop(columns=["Color", "Priority"], errors="ignore").to_excel(
            writer, index=False, sheet_name=sheet_name
        )
        worksheet = writer.sheets[sheet_name]

        apply_excel_formatting(worksheet, df)

        # צביעת עמודת ההערה לפי צבע מהעמודה Color
        for idx, color in enumerate(df.get("Color", []), start=2):
            if pd.isna(color) or color in [None, ""]:
                continue

            color = str(color).strip()

            if color.startswith("#"):
                color = color[1:]

            if len(color) != 6:
                continue

            fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )
            worksheet.cell(row=idx, column=6).fill = fill

    return output.getvalue()


# --- Main function for daily update ---
def get_priority_and_color(comment_text):
    general_patterns = [
        # High
        ("אינה מסגרת חוקית", "High"),
        ("מ.א לא תקין", "High"),
        ("ערך לא חוקי בפלוגתי", "High"),
        ("ערך לא חוקי בגדודי", "High"),
        ("שים לב לנפקדות", "High"),
        ("התייצבות חדשה", "High"),
        ("מסיים שמפ", "High"),
        ("נדרש לסגור שמ\"פ", "High"),
        ("דווח שהתייצב, אך חסר דיווח", "High"),
        ("תאריך התייצבות לא תקין", "High"),
        ("כפילות מ.א", "High"),

        # Medium
        ("לא נמצא בגדודי", "Medium"),
        ("לא נמצא בפלוגתי", "Medium"),
        ("נוסף על בסיס", "Medium"),
        ("עודכן לפי הפלוגתי", "Medium"),
        ("גימלים", "Medium"),

        # Low
        ("ריק בגדודי", "Low"),
        ("פלוגתי ריק", "Low"),
        ("חוסר התאמה", "Low"),
    ]

    for pattern, level in general_patterns:
        if pattern in comment_text:
            return level, COMMENT_PRIORITY[level]

    return "Low", COMMENT_PRIORITY["Low"]


def parse_column_date(col):
    if isinstance(col, (pd.Timestamp, datetime)):
        return pd.Timestamp(col)

    col_str = str(col).strip()

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return pd.to_datetime(col_str, format=fmt)
        except Exception:
            pass

    return pd.to_datetime(col_str, errors="coerce", dayfirst=True)


def compare_and_update_values(df_central, df_dep, run_date, dep_filename):
    comments = []
    current_dep_name = get_department_from_filename(dep_filename)

    shared_cols = []
    run_date_dt = pd.Timestamp(run_date).normalize()

    for col in df_dep.columns:
        if col not in df_central.columns:
            continue

        col_date = parse_column_date(col)

        if pd.notna(col_date) and col_date.normalize() <= run_date_dt:
            shared_cols.append(col)

    central_idx_map = df_central.reset_index().set_index("מ.א")["index"].to_dict()
    dep_idx_map = df_dep.reset_index().set_index("מ.א")["index"].to_dict()

    for _, dep_row in df_dep.iterrows():
        id_number = str(dep_row["מ.א"]).strip()
        if not is_valid_id(id_number):
            continue

        primary_dep = str(dep_row.get("מסגרת ראשית", "")).strip()
        if current_dep_name and primary_dep != current_dep_name:
            comments.append([
                str(dep_row.get("מ.א", "")).strip()
            ] + get_soldier_info(dep_row) + [
                f"חייל לא שייך למסגרת הראשית של הקובץ ({primary_dep}) – לא עודכן"
            ])
            continue

        central_idx = central_idx_map.get(id_number)

        if central_idx is None:
            df_central, comment, _ = handle_missing_in_central(df_central, dep_row)
            if comment:
                comments.append(comment)

            # בגלל שנוספה אולי שורה חדשה - חייבים לרענן מיפוי
            central_idx_map = df_central.reset_index().set_index("מ.א")["index"].to_dict()
            continue

        dep_idx = dep_idx_map.get(id_number)
        if dep_idx is None:
            continue

        for col in shared_cols:
            comparison_comments = compare_and_update_cell(
                df_central, df_dep, central_idx, dep_idx, col, run_date
            )
            comments.extend([[id_number] + get_soldier_info(dep_row) + [msg] for msg in comparison_comments])

    return df_central, df_dep, comments


def find_and_add_missing_rows(df_central, df_dep, dep_filename):
    comments = []

    # --- ניקוי וסידור מ.א ---
    df_central["מ.א"] = clean_id_column(df_central["מ.א"])
    df_dep["מ.א"] = clean_id_column(df_dep["מ.א"])

    # --- בדיקת כפילויות מ.א ---
    comments.extend(check_duplicate_ids(df_central, df_dep))

    # --- בדיקת תקינות מסגרות ראשיות ---
    comments.extend(check_valid_departments(df_central, "גדודי"))
    comments.extend(check_valid_departments(df_dep, "פלוגתי"))

    # --- השוואת מזהים ---
    ids_central = set(df_central["מ.א"])
    ids_dep = set(df_dep["מ.א"])
    common_cols = [col for col in df_central.columns if col in df_dep.columns and col != "מ.א"]

    # --- הוספת חיילים חסרים לגדודי ---
    df_central, added_comments_central = add_missing_from_dep(df_central, df_dep, ids_central, ids_dep, common_cols)
    comments.extend(added_comments_central)

    # --- הוספת חיילים חסרים למחלקתי ---
    df_dep, added_comments_dep = add_missing_from_central(df_dep, df_central, ids_dep, ids_central, common_cols,
                                                          dep_filename)
    comments.extend(added_comments_dep)

    # --- השוואת ערכים בסיסיים בין רשומות משותפות ---
    compare_comments = compare_shared_basic_fields(df_central, df_dep, ids_central & ids_dep)
    comments.extend(compare_comments)

    return df_central, df_dep, comments


def update_shabzak(df_central, df_dep, run_date, dep_filename):
    df_central = df_central.copy()
    df_dep = df_dep.copy()

    df_central = normalize_dataframe(df_central)
    df_dep = normalize_dataframe(df_dep)

    all_comments = []

    # שלב 1: הוספת שורות חסרות
    df_central, df_dep, comments_missing = find_and_add_missing_rows(df_central, df_dep, dep_filename)
    all_comments.extend(comments_missing)

    # שלב 2: השוואה ועדכון ערכים
    df_central, df_dep, comments_updates = compare_and_update_values(df_central, df_dep, run_date, dep_filename)
    all_comments.extend(comments_updates)

    comments_df = pd.DataFrame(
        all_comments,
        columns=["מ.א", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "הערה"]
    )

    if not comments_df.empty:
        comments_df["Priority"], comments_df["Color"] = zip(*comments_df["הערה"].map(get_priority_and_color))
    else:
        comments_df["Priority"] = pd.Series(dtype="object")
        comments_df["Color"] = pd.Series(dtype="object")

    # מחזירה קובץ חדש ומעודכן מחולק לגליונות לפי המקורי
    return {
        "df_central": df_central,
        "comments_df": comments_df,
        "central_xlsx": to_excel_bytes(df_central, "שבצק גדודי"),
        "comments_xlsx": to_colored_excel(comments_df, "הערות")
    }


def update_multiple_departments(df_central, dep_files, run_date):
    df_central = df_central.copy()
    all_comments = []
    updated_departments = []

    for dep_file in dep_files:
        try:
            if dep_file.name.endswith("csv"):
                df_dep = pd.read_csv(dep_file)
            else:
                df_dep = pd.read_excel(dep_file)

            df_dep = normalize_dataframe(df_dep)

            result = update_shabzak(df_central, df_dep, run_date, dep_file.name)
            df_central = result["df_central"]

            comments_df = result["comments_df"].copy()
            if not comments_df.empty:
                all_comments.append(comments_df)

            current_dep_name = get_department_from_filename(dep_file.name)
            if current_dep_name and current_dep_name not in updated_departments:
                updated_departments.append(current_dep_name)
            elif not current_dep_name:
                updated_departments.append(Path(dep_file.name).stem)

        except Exception as e:
            error_df = pd.DataFrame([{
                "מ.א": "",
                "שם פרטי": "",
                "שם משפחה": "",
                "מסגרת ראשית": "",
                "מסגרת משנית": "",
                "הערה": f"שגיאה בעיבוד הקובץ {dep_file.name}: {e}",
                "Priority": "High",
                "Color": COMMENT_PRIORITY["High"],
                "קובץ מקור": dep_file.name
            }])
            all_comments.append(error_df)

    if all_comments:
        comments_df = pd.concat(all_comments, ignore_index=True)
    else:
        comments_df = pd.DataFrame(columns=[
            "מ.א", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית",
            "הערה", "Priority", "Color", "קובץ מקור"
        ])

    return {
        "df_central": df_central,
        "comments_df": comments_df,
        "central_xlsx": to_excel_bytes(df_central, "שבצק גדודי"),
        "comments_xlsx": to_colored_excel(comments_df, "הערות"),
        "updated_departments": updated_departments
    }


################# הוספת מסופחים חדשים #################

def dep_fromRequest(request_dep):
    if pd.isna(request_dep):
        return ""
    request_dep = str(request_dep).strip()
    return requests_TO_DEPARTMENT.get(request_dep, request_dep)


def is_approved_request(status):
    if pd.isna(status):
        return False

    status = str(status).strip()
    return "מאושר" in status and "לא מאושר" not in status


def split_full_name(full_name: str) -> tuple[str, str]:
    if pd.isna(full_name):
        return "", ""

    parts = str(full_name).strip().split()

    if not parts:
        return "", ""

    if len(parts) == 1:
        return parts[0], ""

    special_words = {"אבו", "בן", "בר"}

    for i, word in enumerate(parts):
        if i > 0 and word in special_words:
            first_name = " ".join(parts[:i])
            last_name = " ".join(parts[i:])
            return first_name, last_name

    first_name = " ".join(parts[:-1])
    last_name = parts[-1]
    return first_name, last_name


def build_attached_row(df_central, id_number, first_name, last_name, main_framework, secondary_framework):
    new_row = {col: "" for col in df_central.columns}

    new_row["מ.א"] = id_number
    new_row["שם פרטי"] = first_name
    new_row["שם משפחה"] = last_name
    new_row["מסגרת ראשית"] = main_framework
    new_row["מסגרת משנית"] = secondary_framework
    new_row["מסופח"] = "V"
    new_row['סו"ש'] = "מילואים"

    return new_row


def find_and_add_attached_soldiers(df_central, df_requests):
    comments = []
    added_rows = []

    df_central = df_central.copy()
    df_requests = df_requests.copy()

    df_central["מ.א"] = clean_id_column(df_central["מ.א"])
    df_requests["מ.א. של החייל הדרוש"] = clean_id_column(df_requests["מ.א. של החייל הדרוש"])

    existing_ids = set(df_central["מ.א"])

    for _, request_row in df_requests.iterrows():
        id_number = str(request_row.get("מ.א. של החייל הדרוש", "")).strip()

        first_name, last_name = split_full_name(request_row.get("שם של החייל הדרוש", ""))
        main_framework = dep_fromRequest(request_row.get("פלוגה", ""))
        secondary_framework = format_cell(request_row.get("מחלקה מיועדת", ""))
        status = format_cell(request_row.get("סטטוס קישור", ""))

        soldier_info = [id_number, first_name, last_name, main_framework, secondary_framework]

        if id_number not in existing_ids and is_approved_request(status):
            if not is_valid_id(id_number):
                comments.append(soldier_info + ["מ.א לא תקין"])

            if main_framework not in DEPARTMENTS_LIST:
                comments.append(soldier_info + [f"'{main_framework}' אינה מסגרת חוקית  "])

            new_row = build_attached_row(df_central, id_number, first_name, last_name, main_framework,
                                         secondary_framework)
            df_central = pd.concat([df_central, pd.DataFrame([new_row])], ignore_index=True)
            added_rows.append(new_row)
            existing_ids.add(id_number)

            comments.append(soldier_info + ["בקשת סיפוח מאושרת - החייל נוסף לשבצק הגדודי"])

        df_added = pd.DataFrame(added_rows)
        if df_added.empty:
            df_added = pd.DataFrame(columns=df_central.columns)
        else:
            df_added = df_added.reindex(columns=df_central.columns)

    return df_added, comments


def update_attached_soldiers(df_central, df_requests):
    df_central = df_central.copy()
    df_requests = df_requests.copy()

    all_comments = []

    df_added, comments_added = find_and_add_attached_soldiers(df_central, df_requests)
    all_comments.extend(comments_added)

    comments_df = pd.DataFrame(
        all_comments,
        columns=["מ.א", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "הערה"]
    )

    if not comments_df.empty:
        comments_df["Priority"], comments_df["Color"] = zip(*comments_df["הערה"].map(get_priority_and_color))
    else:
        comments_df["Priority"] = pd.Series(dtype="object")
        comments_df["Color"] = pd.Series(dtype="object")

    return {
        "df_added": df_added,
        "comments_df": comments_df,
        "added_xlsx": to_excel_bytes(df_added, "מסופחים חדשים"),
        "comments_xlsx": to_colored_excel(comments_df, "הערות")
    }


##################################### Platform #################################

# --- Settings ---
st.set_page_config(page_title="השליש האוטומטי", layout="centered")
st.markdown("""
    <style>
    body, .css-18e3th9, .css-1d391kg, .stTextInput, .stButton, .stSelectbox, .stFileUploader {
        direction: rtl;
        text-align: right;
    }
    .stMarkdown, .stText, .stSubheader, .stHeader {
        direction: rtl;
        text-align: right;
    }
    th, td {
        direction: rtl !important;
        text-align: right !important;
    }
    </style>
""", unsafe_allow_html=True)

if "page" not in st.session_state:
    st.session_state.page = "home"

# --- Home ---
if st.session_state.page == "home":
    st.markdown("<h1 style='text-align: center;'>השליש האוטומטי - מגן יהונתן 8552</h1>", unsafe_allow_html=True)

    center_col = st.columns([1, 1, 2, 1, 1])[2]
    with center_col:
        st.image("gdud8552.jpg", width=200)
    #
    # col_center = st.columns([1, 1, 2, 1, 1])[2]
    # with col_center:
    #     if st.button("📆 עדכון שבצ\"ק יומי", use_container_width=True):
    #         go_to("daily_update")
    col_center = st.columns([1, 1, 2, 1, 1])[2]
    with col_center:
        if st.button("📆 עדכון שבצ\"ק יומי", use_container_width=True):
            go_to("daily_update")

        if st.button("💂 עדכון חיילים מסופחים", use_container_width=True):
            go_to("attached_update")

# --- Daily update ---
elif st.session_state.page == "daily_update":
    st.markdown("<h1 style='text-align: center;'>עדכון שבצ\"ק יומי - מגן יהונתן 8552</h1>", unsafe_allow_html=True)

    center_col = st.columns([1, 1, 2, 1, 1])[2]
    with center_col:
        st.image("gdud8552.jpg", width=200)

    has_existing_daily_run = "df_central" in st.session_state and "run_date" in st.session_state

    # אתחול session_state
    if "updated_departments" not in st.session_state:
        st.session_state["updated_departments"] = []

    if "central_filename_base" not in st.session_state:
        st.session_state["central_filename_base"] = None

    central_file = None

    if not has_existing_daily_run:
        central_file = st.file_uploader(
            "**בחר שבצק גדודי**",
            type=["xlsx", "xls", "csv"],
            key="daily_central_file"
        )

        dep_files = st.file_uploader(
            "**בחר קבצי שבצק פלוגתיים**",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            key="daily_dep_files"
        )

        run_date = st.date_input(
            "**בחר תאריך יעד לבדיקה ולעדכון**",
            value=datetime.today().date(),
            format="DD/MM/YYYY",
            key="daily_run_date"
        )

    else:
        saved_run_date = st.session_state["run_date"]

        st.markdown("<h4 style='text-align: right;'>📌 סטטוס עדכון יומי</h4>", unsafe_allow_html=True)
        st.success(f"מעודכן לתאריך יעד: {saved_run_date.strftime('%d/%m/%Y')}")

        updated_deps = st.session_state.get("updated_departments", [])
        if updated_deps:
            st.markdown("**פלוגות שעודכנו עד כה:**")

            num_cols = 3
            max_rows = (len(updated_deps) + num_cols - 1) // num_cols

            for row_idx in range(max_rows):
                cols = st.columns(num_cols)

                for col_idx in range(num_cols):
                    i = row_idx * num_cols + col_idx

                    if i < len(updated_deps):
                        dep = updated_deps[i]

                        with cols[col_idx]:
                            st.markdown(
                                f"""
                                <div style="
                                    background-color:#d1f7d6;
                                    border:1px solid #28a745;
                                    border-radius:12px;
                                    padding:6px;
                                    text-align:center;
                                    font-weight:600;
                                    color:#155724;
                                ">
                                    ✔ {dep}
                                </div>
                                """,
                                unsafe_allow_html=True
                            )

        st.markdown("<br>", unsafe_allow_html=True)

        dep_files = st.file_uploader(
            "**בחר שבצקי פלוגה נוספים**",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            key="daily_dep_files_next"
        )

        run_date = saved_run_date

    if st.button("🚀 עדכן שבצק"):
        if not dep_files:
            st.error("יש להעלות לפחות קובץ שבצק פלוגתי אחד.")
        elif not has_existing_daily_run and not central_file:
            st.error("יש להעלות גם קובץ שבצק גדודי בריצה הראשונה.")
        else:
            if not has_existing_daily_run:
                try:
                    if central_file.name.endswith("csv"):
                        df_central = pd.read_csv(central_file)
                    else:
                        df_central = pd.read_excel(central_file)

                    df_central = normalize_dataframe(df_central)
                    st.session_state["central_filename_base"] = Path(central_file.name).stem

                except Exception as e:
                    st.error(f"שגיאה בטעינת שבצק גדודי: {e}")
                    st.stop()
            else:
                df_central = st.session_state["df_central"]

            st.success("✅ הקבצים נטענו בהצלחה. מוכן להריץ בדיקות.")

            st.session_state["run_date"] = run_date
            with st.spinner("מעדכן שבצ״ק, זה יקח כמה רגעים..."):
                result = update_multiple_departments(df_central, dep_files, run_date)
            # result = update_multiple_departments(df_central, dep_files, run_date)

            st.session_state["download_ready"] = True
            st.session_state["central_xlsx"] = result["central_xlsx"]
            st.session_state["comments_xlsx"] = result["comments_xlsx"]
            st.session_state["comments_df"] = result["comments_df"]
            st.session_state["df_central"] = result["df_central"]

            for dep_name in result["updated_departments"]:
                if dep_name not in st.session_state["updated_departments"]:
                    st.session_state["updated_departments"].append(dep_name)

            if "all_comments_df" in st.session_state:
                st.session_state["all_comments_df"] = pd.concat(
                    [st.session_state["all_comments_df"], result["comments_df"]],
                    ignore_index=True
                )
            else:
                st.session_state["all_comments_df"] = result["comments_df"]

    if st.session_state.get("download_ready"):
        updated_deps = st.session_state.get("updated_departments", [])

        if updated_deps:
            st.markdown("<h3 style='text-align: right;'>השבצ\"ק עודכן עבור הפלוגות הבאות</h3>", unsafe_allow_html=True)

            num_cols = 3
            max_rows = (len(updated_deps) + num_cols - 1) // num_cols

            for row_idx in range(max_rows):
                cols = st.columns(num_cols)

                for col_idx in range(num_cols):
                    i = row_idx * num_cols + col_idx

                    if i < len(updated_deps):
                        dep = updated_deps[i]

                        with cols[col_idx]:
                            st.markdown(
                                f"""
                                <div style="
                                    background-color:#d1f7d6;
                                    border:1px solid #28a745;
                                    border-radius:12px;
                                    padding:6px;
                                    text-align:center;
                                    font-weight:600;
                                    color:#155724;
                                ">
                                    ✔ {dep}
                                </div>
                                """,
                                unsafe_allow_html=True
                            )

        st.markdown("<h3 style='text-align: right;'>📥 הורדת קבצים מעודכנים</h3>", unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        central_filename_base = st.session_state.get("central_filename_base", "שבצק_גדודי")
        central_filename = central_filename_base + "_מעודכן.xlsx"

        with col1:
            st.download_button(
                label="📁 הורד שבצ\"ק גדודי",
                data=st.session_state["central_xlsx"],
                file_name=central_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col2:
            st.download_button(
                label="📝 הורד דוח הערות",
                data=to_colored_excel(st.session_state["all_comments_df"], "הערות"),
                file_name="דוח_הערות.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("<h3 style='text-align: right;'>📋 תצוגת דוח הערות</h3>", unsafe_allow_html=True)
        render_comments_table(st.session_state["all_comments_df"])

        if st.button("עדכון פלוגות נוספות"):
            st.session_state.pop("download_ready", None)
            st.session_state.pop("central_xlsx", None)
            st.session_state.pop("comments_xlsx", None)
            go_to("daily_update")

    if st.button("⬅️ חזרה למסך הראשי"):
        st.session_state.pop("download_ready", None)
        st.session_state.pop("central_xlsx", None)
        st.session_state.pop("comments_xlsx", None)
        st.session_state.pop("comments_df", None)
        st.session_state.pop("df_central", None)
        st.session_state.pop("run_date", None)
        st.session_state.pop("all_comments_df", None)
        st.session_state.pop("updated_departments", None)
        st.session_state.pop("central_filename_base", None)

        go_to("home")

elif st.session_state.page == "attached_update":
    st.markdown("<h1 style='text-align: center;'>עדכון חיילים מסופחים בשבצק - מגן יהונתן 8552</h1>",
                unsafe_allow_html=True)

    center_col = st.columns([1, 1, 2, 1, 1])[2]
    with center_col:
        st.image("gdud8552.jpg", width=200)

    central_file = st.file_uploader("בחר שבצק גדודי", type=["xlsx", "xls", "csv"], key="attached_central_file")
    requests_file = st.file_uploader("בחר קובץ בקשות סיפוח", type=["xlsx", "xls", "csv"], key="attached_requests_file")

    if st.button("💂 עדכן מסופחים"):
        if not central_file or not requests_file:
            st.error("יש להעלות גם קובץ שבצק גדודי וגם קובץ בקשות סיפוח.")
        else:
            try:
                if central_file.name.endswith("csv"):
                    df_central = pd.read_csv(central_file)
                else:
                    df_central = pd.read_excel(central_file)
                    df_central = normalize_dataframe(df_central)

            except Exception as e:
                st.error(f"שגיאה בטעינת השבצק הגדודי: {e}")
                st.stop()

            try:
                if requests_file.name.endswith("csv"):
                    df_requests = pd.read_csv(requests_file)
                else:
                    df_requests = pd.read_excel(requests_file)
                    df_requests = normalize_dataframe(df_requests)

            except Exception as e:
                st.error(f"שגיאה בטעינת קובץ בקשות הסיפוח: {e}")
                st.stop()

            st.success("✅ הקבצים נטענו בהצלחה. מוכן לעדכון מסופחים.")

            result = update_attached_soldiers(df_central, df_requests)

            st.session_state["attached_download_ready"] = True
            st.session_state["attached_added_xlsx"] = result["added_xlsx"]
            st.session_state["attached_comments_df"] = result["comments_df"]
            st.session_state["attached_comments_xlsx"] = result["comments_xlsx"]
            st.session_state["attached_df_added"] = result["df_added"]

    if st.session_state.get("attached_download_ready"):
        st.markdown("<h3 style='text-align: right;'>📥 הורדת קבצים מעודכנים</h3>", unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        central_filename = Path(central_file.name).stem + "_מעודכן.xlsx"

        with col1:
            st.download_button(
                label="📁 הורד דוח מסופחים חדשים לשבצ\"ק",
                data=st.session_state["attached_added_xlsx"],
                file_name="מסופחים_חדשים_להוספה.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col2:
            st.download_button(
                label="📝 הורד דוח הערות",
                data=st.session_state["attached_comments_xlsx"],
                file_name="דוח_הערות_מסופחים.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("<h3 style='text-align: right;'>📋 תצוגת דוח הערות</h3>", unsafe_allow_html=True)
        render_comments_table(st.session_state["attached_comments_df"])

    if st.button("⬅️ חזרה למסך הראשי", key="back_from_attached"):
        st.session_state.pop("attached_download_ready", None)
        st.session_state.pop("attached_added_xlsx", None)
        st.session_state.pop("attached_comments_df", None)
        st.session_state.pop("attached_comments_xlsx", None)
        st.session_state.pop("attached_df_added", None)

        go_to("home")
