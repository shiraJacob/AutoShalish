import openpyxl
import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
from datetime import datetime
import re
from openpyxl.styles import PatternFill, Border, Side


######################### Constant variables #########################

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
DEPARTMENTS_LIST = ["פלוגה מבצעית א'", "פלוגה מבצעית ב'", "פלוגה מבצעית ג'", "פלוגה מסייעת", "מפג\"ד", "פלס\"ם", "אג\"מ"]
valid_roles = ["מ\"פ ומעלה", "לוחם", "יתר", "נהג בט\"ש", "מט\"ב", "נהג משא", "חובש"]
valid_sosh = ["סדיר", "מילואים"]
valid_status = ["V", "התייצב ושוחרר", "שוחרר"]

unintended_values = ["נ", "0"]

valid_AGAM = ["ח", "ג", "נ", "י", "0", "עתניאל" ,"1", "ימי התארגנות"]
valid_PALSAM = ["ח", "ג", "נ", "י", "0", "עתניאל", "אדוריים", "בית חגי", "סוסיא", "חוות מעון", "1", "ימי התארגנות"]
valid_MAFGAD = ["ח", "ג", "נ", "י", "0", "עתניאל", "1", "ימי התארגנות"]
valid_A = ["ח", "ג", "נ", "י", "0", "אדוריים", "חורסה", "כרם נגוהות", "1", "ימי התארגנות"]
valid_B = ["ח", "ג", "נ", "י", "0", "סוסיא", "720", "1", "חוות מעון", "ימי התארגנות"]
valid_C = ["ח", "ג", "נ", "י", "0", "710", "620","בית חגי", "1", "ימי התארגנות"]
valid_D = ["ח", "ג", "נ", "י", "0", "עתניאל", "630", "1", "ימי התארגנות"]

######################### Mapping functions #########################

FILENAME_TO_DEPARTMENT = {
    "שבצק פלוגה א'": "פלוגה מבצעית א'",
    "שבצק פלוגה ב'": "פלוגה מבצעית ב'",
    "שבצק פלוגה ג'": "פלוגה מבצעית ג'",
    "שבצק פלוגה מסייעת": "פלוגה מסייעת",
    "שבצק אג\"מ": "אג\"מ",
    "שבצק פלס\"ם": "פלס\"ם",
    "שבצק מפג\"ד": "מפג\"ד",
}

UNIT_VALID_MAP = {
    "אג\"מ": valid_AGAM,
    "אג_מ": valid_AGAM,
    "אגמ": valid_AGAM,
    "פלסם": valid_PALSAM,
    "פלס\"ם": valid_PALSAM,
    "פלס_ם": valid_PALSAM,
    "מפגד": valid_MAFGAD,
    "מפג_ד": valid_MAFGAD,
    "מפג\"ד": valid_MAFGAD,
    "פלוגה א'": valid_A,
    "פלוגה ב'": valid_B,
    "פלוגה ג'": valid_C,
    "פלוגה מסייעת": valid_D,
}

dep_to_central_map = {
    "ח": "ח",
    "י": "י"
   # "ג": "בית",
   # "נ": "בית",
   # "0": "בית"
}

central_to_dep_map = {
    "ח": "ח",
    "ימי התארגנות" : "ימי התארגנות"
}

dep_map = {
    "מפקדת היחידה": "מפג\"ד",
    "פלוגה מבצעית א'": "פלוגה א'",
    "פלוגה מבצעית ב'": "פלוגה ב'",
    "פלוגה מבצעית ג'": "פלוגה ג'",
    "פלוגה מבצעית ד'": "פלוגה מסייעת'"
}

COMMENT_PRIORITY = {
    "High": "#f8d7da",    # pastel red
    "Medium": "#fff3cd",  # pastel yellow
    "Low": None
}

############################# Functions #####################################

################# General #################

def go_to(page_name):
    st.session_state.page = page_name

def format_cell(val):
    if pd.isna(val):
        return ""
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%d/%m/%y")
    return str(val).strip() if not isinstance(val, float) else str(int(val)).strip()

def get_valid_values_by_filename(filename: str):
    filename = filename.lower()
    for key, valid_list in UNIT_VALID_MAP.items():
        if key.replace('"', '').lower() in filename:
            return valid_list
    return valid_values
def render_comments_table(df):
    styled_rows = []
    for _, row in df.iterrows():
        color = row.get("Color")
        bg = f'background-color:{color};' if color else ''
        styled_rows.append(
            f"<tr style='{bg}'>" +
            "".join(f"<td style='white-space: nowrap; padding: 6px 10px;'>{row[col]}</td>"
                    for col in ["מ.א.", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "הערה"]) +
            "</tr>"
        )

    html = f"""
    <div style='max-height: 600px; overflow-y: auto; overflow-x: auto; border: 1px solid #ccc; border-radius: 6px;'>
        <table style='border-collapse: collapse; width: 100%; direction: rtl; font-size: 14px;'>
            <thead>
                <tr style='background-color: #f0f0f0; text-align: right;'>
                    <th style='padding: 8px;'>מ.א.</th>
                    <th style='padding: 8px;'>שם פרטי</th>
                    <th style='padding: 8px;'>שם משפחה</th>
                    <th style='padding: 8px;'>מסגרת ראשית</th>
                    <th style='padding: 8px;'>מסגרת משנית</th>
                    <th style='padding: 8px;'>הערה</th>
                </tr>
            </thead>
            <tbody>
                {''.join(styled_rows)}
            </tbody>
        </table>
    </div>
    """

    st.markdown(html, unsafe_allow_html=True)




################# Daily Update #################

def is_valid_id(id_number):
    if pd.isna(id_number):
        return False
    id_str = str(id_number).strip()
    return id_str.isdigit() and len(id_str) == 7

# ---- Compare two shabzaks ---
def get_soldier_info(row):
    return [
        str(row.get("שם פרטי", "")).strip(),
        str(row.get("שם משפחה", "")).strip(),
        str(row.get("מסגרת ראשית", "")).strip(),
        str(row.get("מסגרת משנית", "")).strip()
    ]

def is_valid_value_for_column(col: str, val: str) -> bool:
    val = str(val).strip()
    if col == "תפקיד":
        return val in valid_roles
    elif col == "סו\"ש":
        return val in valid_sosh
    elif col == "התייצב":
        return val in valid_status
    return True
def normalize_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%d/%m/%y")  # כבר אובייקט תאריך
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%y").strftime("%d/%m/%y")
    except Exception:
        return None
def handle_missing_in_central(df_central, dep_row):
    id_number = str(dep_row["מ.א."]).strip()
    central_match = df_central[df_central["מ.א."] == id_number]

    if central_match.empty:
        if not (id_number.isdigit() and len(id_number) == 7):
            return df_central, [
                id_number,
                dep_row.get("שם פרטי", ""),
                dep_row.get("שם משפחה", ""),
                dep_row.get("מסגרת ראשית", ""),
                dep_row.get("מסגרת משנית", ""),
                "מ.א. לא תקין, לא התבצע עדכון"
            ], None

        common_cols = [col for col in df_central.columns if col in dep_row and col != "מ.א."]
        new_row = {col: dep_row[col] for col in common_cols}
        new_row["מ.א."] = id_number
        df_central = pd.concat([df_central, pd.DataFrame([new_row])], ignore_index=True)

        return df_central, [
            id_number,
            dep_row.get("שם פרטי", ""),
            dep_row.get("שם משפחה", ""),
            dep_row.get("מסגרת ראשית", ""),
            dep_row.get("מסגרת משנית", ""),
            "לא נמצא במרוכז – נוסף על בסיס הפלוגתי"
        ], None

    return df_central, None, central_match.index[0]
def analyze_suspicious_cases(df_central, df_dep, central_idx, dep_idx, col, central_val, dep_val):
    suspicious_comments = []
    dep_status = format_cell(df_dep.at[dep_idx, "התייצב"])
    central_status = format_cell(df_central.at[central_idx, "התייצב"])
    m_col = format_cell(col)
    if dep_val == "ג" or central_val == "ג":
        if dep_status != "V" or central_status != "V":
            suspicious_comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', לא דווח שהתייצב – לבדוק גימלים")
        else:
            suspicious_comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', התייצב ויצא לגימלים – לבדוק אישור רופא")

    if dep_val == "נ" or central_val == "נ":
        col_date = pd.to_datetime(col, errors="coerce")
        today = pd.Timestamp.today().normalize()

        if pd.notna(col_date) and col_date.normalize() == today:
            suspicious_comments.append(
                f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', שים לב לנפקדות"
            )

    if dep_val in ["0", "ימי התארגנות"] or central_val in ["0", "ימי התארגנות"]:
        if dep_status == "V" or central_status == "V":
            try:
                enlist_raw = df_central.at[central_idx, "תאריך התייצבות"]
                enlist_str = normalize_date(enlist_raw)

                m_col_date = datetime.strptime(m_col.strip(), "%d/%m/%y")
                enlist_date = datetime.strptime(enlist_str, "%d/%m/%y")

            except Exception as e:
                suspicious_comments.append(
                    f"{m_col}: תאריך התייצבות לא תקין – לא בוצע עדכון סטטוס (שגיאה: {e})"
                )
                return suspicious_comments

            # אם התאריך של 0/ימי התארגנות הוא אחרי או שווה לתאריך התייצבות → נחשב סיום שמפ
            if m_col_date >= enlist_date:
                df_central.at[central_idx, "התייצב"] = "התייצב ושוחרר"
                df_dep.at[dep_idx, "התייצב"] = "התייצב ושוחרר"
                suspicious_comments.append(
                    f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', סיום שמפ אתמול – ודא הזדכות על ציוד"
                )

    return suspicious_comments
def compare_and_update_cell(df_central, df_dep, central_idx, dep_idx, col):
    comments = []
    central_val = format_cell(df_central.at[central_idx, col])
    dep_val = format_cell(df_dep.at[dep_idx, col])

    status_central = format_cell(df_central.at[central_idx, "התייצב"])
    status_dep = format_cell(df_dep.at[dep_idx, "התייצב"])

    is_arrived = status_central == "V" or status_dep == "V"

    is_central_empty = central_val in ["", "nan"]
    is_dep_empty = dep_val in ["", "nan"]

    # --- בדיקות חשודות ---
    comments.extend(analyze_suspicious_cases(df_central, df_dep, central_idx, dep_idx, col, central_val, dep_val))

    # --- השוואה ועדכון ---
    m_col = format_cell(col)
    if not is_dep_empty and dep_val not in valid_values:
        comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', ערך לא חוקי בפלוגתי – לא בוצע עדכון")

    elif is_central_empty and not is_dep_empty:
        if not is_arrived and dep_val not in ["ח", "ג", "נ", "י", "0", "ימי התארגנות"]:
            comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}  : התייצבות חדשה – לדאוג לפתוח שמפ ולעדכן תאריך התייצבות ולוודא מעבר בשלישות")
            df_central.at[central_idx, "התייצב"] = "V"
            df_dep.at[dep_idx, "התייצב"] = "V"
            df_central.at[central_idx, "תאריך התייצבות"] = m_col
            df_dep.at[dep_idx, "תאריך התייצבות"] = m_col
        df_central.at[central_idx, col] = dep_to_central_map.get(dep_val, dep_val)

    elif not is_central_empty and not is_dep_empty and not (
            central_val == dep_to_central_map.get(dep_val, dep_val) or
            dep_val == central_to_dep_map.get(central_val, central_val)
    ):
        if central_val == "ימי התארגנות":
            df_dep.at[dep_idx, col] = central_to_dep_map.get(central_val, central_val)
            # comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', עודכן לפי המרוכז")
        else:
            df_central.at[central_idx, col] = dep_to_central_map.get(dep_val, dep_val)
            comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', עודכן לפי הפלוגתי")

    elif not is_central_empty and is_dep_empty:
        if central_val not in valid_values:
            comments.append(f"{m_col}: מרוכז='{central_val}', פלוגתי='{dep_val}', ערך לא חוקי במרוכז – לא בוצע עדכון")
        else:
            df_dep.at[dep_idx, col] = central_to_dep_map.get(central_val, central_val)
            if not is_arrived and central_val not in ["ח", "ג", "נ", "י", "0", "ימי התארגנות"]:
                df_central.at[central_idx, "התייצב"] = "V"
                df_dep.at[dep_idx, "התייצב"] = "V"
                df_central.at[central_idx, "תאריך התייצבות"] = m_col
                df_dep.at[dep_idx, "תאריך התייצבות"] = m_col
                comments.append(f"{m_col}: הושלם מהמרוכז='{central_val}',   : התייצבות חדשה – לדאוג לפתוח שמפ ולעדכן תאריך התייצבות ולוודא מעבר בשלישות")
            comments.append(f"{m_col}: פלוגתי ריק, הושלם מהמרוכז='{central_val}'")


    elif is_central_empty and is_dep_empty and is_arrived:
        enlistment_date_str = df_central.at[central_idx, "תאריך התייצבות"]
        try:
            # פירוש תאריכים בפורמט יום/חודש/שנה
            enlistment_date = pd.to_datetime(enlistment_date_str, errors="coerce", dayfirst=True)
            col_date = pd.to_datetime(col, errors="coerce", dayfirst=True)
            today = pd.Timestamp.today().normalize().date()
            if pd.notna(col_date) and pd.notna(enlistment_date):
                col_date = col_date.date()
                enlistment_date = enlistment_date.date()
                if enlistment_date <= col_date <= today:
                    comments.append(f"{m_col}: דווח שהתייצב, אך חסר דיווח")
            else:
                comments.append(f"{m_col}: תאריך התייצבות לא תקין")
        except Exception as e:
            print("❌ ERROR in התייצבות check:", e)

    return comments



#########################################

def compare_shared_basic_fields(df_central, df_dep, common_ids):
    comments = []
    fields_to_check = ["שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "התייצב", "סו\"ש", "תפקיד"]
    for id_number in common_ids:
        if not is_valid_id(id_number):
            continue

        central_idx = df_central.index[df_central["מ.א."] == id_number][0]
        dep_idx = df_dep.index[df_dep["מ.א."] == id_number][0]

        for col in fields_to_check:
            if col not in df_central.columns or col not in df_dep.columns:
                continue

            central_val = format_cell(df_central.at[central_idx, col])
            dep_val = format_cell(df_dep.at[dep_idx, col])

            central_empty = central_val in ["", "nan", "None", "NaT"]
            dep_empty = dep_val in ["", "nan", "None", "NaT"]

            # מרוכז ריק, פלוגתי מלא - נעדכן לפי פלוגתי רק אם ערך חוקי
            if central_empty and not dep_empty:
                if is_valid_value_for_column(col, dep_val):
                    df_central.at[central_idx, col] = dep_to_central_map.get(dep_val, dep_val)
                    comments.append([
                        id_number,
                        df_dep.at[dep_idx, "שם פרטי"],
                        df_dep.at[dep_idx, "שם משפחה"],
                        df_dep.at[dep_idx, "מסגרת ראשית"],
                        df_dep.at[dep_idx, "מסגרת משנית"],
                        f"{col}: ריק במרוכז, הושלם לערך '{dep_val}'"
                    ])
                else:
                    comments.append([
                        id_number,
                        df_dep.at[dep_idx, "שם פרטי"],
                        df_dep.at[dep_idx, "שם משפחה"],
                        df_dep.at[dep_idx, "מסגרת ראשית"],
                        df_dep.at[dep_idx, "מסגרת משנית"],
                        f"{col}: ריק במרוכז, ערך לא חוקי בפלוגתי ('{dep_val}') – לא הושלם למרוכז"
                    ])

            # פלוגתי ריק, מרוכז מלא - נעדכן לפי מרוכז רק אם ערך חוקי
            elif dep_empty and not central_empty:
                if is_valid_value_for_column(col, central_val):
                    df_dep.at[dep_idx, col] = central_to_dep_map.get(central_val, central_val)
                    #comments.append([
                      #  id_number,
                     #   df_dep.at[dep_idx, "שם פרטי"],
                    #  df_dep.at[dep_idx, "שם משפחה"],
                     #   df_dep.at[dep_idx, "מסגרת ראשית"],
                     #   df_dep.at[dep_idx, "מסגרת משנית"],
                    #    f"{col}: ריק בפלוגתי – הושלם לפי המרוכז ('{central_val}')"
                   # ])
                else:
                    comments.append([
                        id_number,
                        df_dep.at[dep_idx, "שם פרטי"],
                        df_dep.at[dep_idx, "שם משפחה"],
                        df_dep.at[dep_idx, "מסגרת ראשית"],
                        df_dep.at[dep_idx, "מסגרת משנית"],
                        f"{col}: פלוגתי ריק, ערך לא חוקי במרוכז ('{central_val}') – לא הושלם לפלוגתי"
                    ])

            # פלוגתי ומרוכז מלאים אבל שונים זה מזה - נעדכן לפי מרוכז את כל העמודות חוץ ממסגרת משנית שאותה נעדכן לפי פלוגתי
            elif not central_empty and not dep_empty and not (
                central_val == dep_to_central_map.get(dep_val, dep_val) or
                dep_val == central_to_dep_map.get(central_val, central_val)
            ):
                if is_valid_value_for_column(col, dep_val):
                    if col == "מסגרת משנית":
                        df_central.at[central_idx, col] = dep_to_central_map.get(dep_val, dep_val)
                        comments.append([
                            id_number,
                            df_dep.at[dep_idx, "שם פרטי"],
                            df_dep.at[dep_idx, "שם משפחה"],
                            df_dep.at[dep_idx, "מסגרת ראשית"],
                            df_dep.at[dep_idx, "מסגרת משנית"],
                            f"{col}: חוסר התאמה – עודכן לפי הפלוגתי ('{dep_val}')"
                        ])
                    else:
                        df_dep.at[dep_idx, col] = central_to_dep_map.get(central_val, central_val)
                       # comments.append([
                        #    id_number,
                        #    df_dep.at[dep_idx, "שם פרטי"],
                        #    df_dep.at[dep_idx, "שם משפחה"],
                        #    df_dep.at[dep_idx, "מסגרת ראשית"],
                        #    df_dep.at[dep_idx, "מסגרת משנית"],
                        #    f"{col}: חוסר התאמה – עודכן לפי המרוכז ('{dep_val}')"
                       # ])
                else:
                    comments.append([
                        id_number,
                        df_dep.at[dep_idx, "שם פרטי"],
                        df_dep.at[dep_idx, "שם משפחה"],
                        df_dep.at[dep_idx, "מסגרת ראשית"],
                        df_dep.at[dep_idx, "מסגרת משנית"],
                        f"{col}: אי תאימות - ערך לא חוקי {central_val}  ('{dep_val}') – לא עודכן "
                    ])
    return comments

def get_department_from_filename(filename):
    filename = Path(filename).stem.strip()
    for key, dept in FILENAME_TO_DEPARTMENT.items():
        if key in filename:
            return dept
    return None  # לא מזוהה

def clean_id_column(series):
    return series.apply(lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() else str(x).strip())

def check_duplicate_ids(df_central, df_dep):
    comments = []
    duplicated_central = df_central[df_central.duplicated("מ.א.", keep=False)]
    duplicated_dep = df_dep[df_dep.duplicated("מ.א.", keep=False)]

    for _, row in duplicated_central.iterrows():
        comments.append([
            row["מ.א."], row.get("שם פרטי", ""), row.get("שם משפחה", ""),
            row.get("מסגרת ראשית", ""), row.get("מסגרת משנית", ""),
            "כפילות מ.א. בקובץ מרוכז – נדרש בירור"
        ])

    for _, row in duplicated_dep.iterrows():
        comments.append([
            row["מ.א."], row.get("שם פרטי", ""), row.get("שם משפחה", ""),
            row.get("מסגרת ראשית", ""), row.get("מסגרת משנית", ""),
            "כפילות מ.א. בקובץ פלוגתי – נדרש בירור"
        ])

    return comments

def check_valid_departments(df, source_name):
    comments = []

    for idx, row in df.iterrows():
        raw_val = str(row.get("מסגרת ראשית", "")).strip()

        if raw_val in DEPARTMENTS_LIST:
            continue

        elif raw_val in dep_map:
            df.at[idx, "מסגרת ראשית"] = dep_map[raw_val]
        else:
            comments.append([
                row.get("מ.א.", ""),
                row.get("שם פרטי", ""),
                row.get("שם משפחה", ""),
                raw_val,
                row.get("מסגרת משנית", ""),
                f"'{raw_val}' אינה מסגרת חוקית ({source_name})"
            ])

    return comments

def add_missing_from_dep(df_central, df_dep, ids_central, ids_dep, common_cols):
    comments = []
    missing_in_central = ids_dep - ids_central

    for id_number in missing_in_central:
        row = df_dep[df_dep["מ.א."] == id_number].iloc[0]

        if not is_valid_id(id_number):
            comments.append([
                id_number, row.get("שם פרטי", ""), row.get("שם משפחה", ""),
                row.get("מסגרת ראשית", ""), row.get("מסגרת משנית", ""),
                "מ.א. לא תקין"
            ])
            continue
        new_row = {col: row[col] for col in common_cols}
        new_row["מ.א."] = id_number
        df_central = pd.concat([df_central, pd.DataFrame([new_row])], ignore_index=True)
        comments.append([
            id_number, row.get("שם פרטי", ""), row.get("שם משפחה", ""),
            row.get("מסגרת ראשית", ""), row.get("מסגרת משנית", ""),
            "לא נמצא במרוכז – נוסף על בסיס הפלוגתי"
        ])

    return df_central, comments

def add_missing_from_central(df_dep, df_central, ids_dep, ids_central, common_cols):
    comments = []
    missing_in_dep = ids_central - ids_dep

    current_dep_name = get_department_from_filename(dep_file.name)

    for id_number in missing_in_dep:
        row = df_central[df_central["מ.א."] == id_number].iloc[0]
        if not is_valid_id(id_number):
            comments.append([
                id_number, row.get("שם פרטי", ""), row.get("שם משפחה", ""),
                row.get("מסגרת ראשית", ""), row.get("מסגרת משנית", ""),
                "מ.א. לא תקין"
            ])
            continue
        unit = str(row.get("מסגרת ראשית", "")).strip()
        enlist_date = str(row.get("התייצב", "")).strip()

        if current_dep_name and unit == current_dep_name and enlist_date not in ["שוחרר", "התייצב ושוחרר"]:
            new_row = {col: row[col] for col in common_cols}
            new_row["מ.א."] = id_number
            df_dep = pd.concat([df_dep, pd.DataFrame([new_row])], ignore_index=True)
            comments.append([
                id_number, row.get("שם פרטי", ""), row.get("שם משפחה", ""),
                row.get("מסגרת ראשית", ""), row.get("מסגרת משנית", ""),
                "לא נמצא בפלוגתי – נוסף על בסיס המרוכז"
            ])

    return df_dep, comments


################# Excel #################
def apply_excel_formatting(worksheet, df):
    worksheet.sheet_view.rightToLeft = True  # ✅ יישור מימין לשמאל

    # הגדרות צבעים וגבולות
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # כחול בהיר
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    date_format = "DD/MM/YY"

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                                      min_col=1, max_col=worksheet.max_column), start=1):
        for col_idx, cell in enumerate(row, start=1):
            # צבע רקע לשורת כותרת
            if row_idx == 1:
                cell.fill = header_fill
            else:
                cell.fill = PatternFill(fill_type=None)  # רקע לבן / שקוף

            # גבולות
            cell.border = border

            # פורמט תאריך לעמודות תאריך
            col_letter = worksheet.cell(row=1, column=col_idx).value
            if col_letter and pd.api.types.is_datetime64_any_dtype(df[col_letter]):
                if row_idx > 1:
                    cell.number_format = date_format
def merge_all_sheets(uploaded_file):
    try:
        # טען את כל הגיליונות כ־DataFrames
        all_sheets = pd.read_excel(uploaded_file, sheet_name=None)  # מחזיר dict: sheet_name -> DataFrame

        merged_df = pd.DataFrame()

        for sheet_name, df in all_sheets.items():
            df = df.copy()
            if not df.empty:
                df["sheet"] = sheet_name
                merged_df = pd.concat([merged_df, df], ignore_index=True)

        return merged_df

    except Exception as e:
        st.error(f"שגיאה באיחוד הגיליונות: {e}")
        return None
def split_to_sheets(df):
    df = df.copy()

    if "sheet" not in df.columns or "מסגרת משנית" not in df.columns:
        raise ValueError("העמודות 'sheet' ו-'מסגרת משנית' חייבות להופיע בטבלה")

    existing_sheets = set(df["sheet"].dropna().astype(str).str.strip().unique())

    for idx, row in df[df["sheet"].isna()].iterrows():
        alt_val = str(row["מסגרת משנית"]).strip()

        if alt_val in existing_sheets:
            df.at[idx, "sheet"] = alt_val
        else:
            df.at[idx, "sheet"] = "נוספים"

    result = {}
    for sheet_name, group in df.groupby("sheet"):
        result[sheet_name] = group.drop(columns=["sheet"])  # הסרה של העמודה מהתוצאה אם רוצים

    return result
def to_excel_bytes(df, sheet_name):
    df = df.copy()
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        apply_excel_formatting(worksheet, df)  # ✅ עיצוב מלא
    return output.getvalue()
def dict_to_excel_bytes(sheets_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df = df.copy()
            safe_sheet_name = str(sheet_name)[:31]
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
            worksheet = writer.sheets[safe_sheet_name]
            apply_excel_formatting(worksheet, df)  # ✅ עיצוב מלא
    return output.getvalue()
def to_colored_excel(df, sheet_name):
    df = df.copy()
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.drop(columns=["Color", "Priority"], errors="ignore").to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        apply_excel_formatting(worksheet, df)

        # שמור את צבע הרקע של הערות (עמודת "Color")
        for idx, color in enumerate(df.get("Color", []), start=2):
            if color:
                fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type="solid")
                worksheet.cell(row=idx, column=6).fill = fill

    return output.getvalue()


# --- Main function for daily update ---
def get_priority_and_color(comment_text):
    general_patterns = [
        # High
        ("אינה מסגרת חוקית", "High"),
        ("מ.א. לא תקין", "High"),
        ("ערך לא חוקי בפלוגתי", "High"),
        ("ערך לא חוקי במרוכז", "High"),
        ("שים לב לנפקדות", "High"),
        ("לא דווח שהתייצב", "High"),
        ("התייצבות חדשה", "High"),
        ("סיום שמפ", "High"),
        ("דווח שהתייצב, אך חסר דיווח", "High"),
        ("כפילות מ.א.", "High"),

        # Medium
        ("לא נמצא במרוכז – נוסף", "Medium"),
        ("לא נמצא בפלוגתי – נוסף", "Medium"),
        ("נוסף על בסיס", "Medium"),
        ("עודכן לפי הפלוגתי", "Medium"),
        ("התייצב ויצא לגימלים", "Medium"),

        # Low
        ("ריק במרוכז", "Low"),
        ("ריק בפלוגתי", "Low"),
        ("עודכן לפי המרוכז", "Low"),
        ("פלוגתי ריק", "Low"),
    ]

    for pattern, level in general_patterns:
        if pattern in comment_text:
            return level, COMMENT_PRIORITY[level]

    return "Low", COMMENT_PRIORITY["Low"]
def compare_and_update_values(df_central, df_dep):
    comments = []
    current_dep_name = get_department_from_filename(dep_file.name)

    shared_cols = [
        col for col in df_dep.columns
        if col in df_central.columns
           and pd.notna(pd.to_datetime(str(col), errors="coerce", dayfirst=True))
    ]

    for _, dep_row in df_dep.iterrows():
        id_number = str(dep_row["מ.א."]).strip()
        if not is_valid_id(id_number):
            continue

        primary_dep = str(dep_row.get("מסגרת ראשית", "")).strip()
        if current_dep_name and primary_dep != current_dep_name:
            comments.append([
                str(dep_row.get("מ.א.", "")).strip(),
                dep_row.get("שם פרטי", ""),
                dep_row.get("שם משפחה", ""),
                primary_dep,
                dep_row.get("מסגרת משנית", ""),
                f"חייל לא שייך למסגרת הראשית של הקובץ ({primary_dep}) – לא עודכן, "
            ])
            continue

        df_central, comment, central_idx = handle_missing_in_central(df_central, dep_row)
        if comment:
            comments.append(comment)
            continue  # לא קיים במרוכז, נוצר – נעבור לחייל הבא

        dep_idx = df_dep[df_dep["מ.א."] == id_number].index[0]

        for col in shared_cols:
            comparison_comments = compare_and_update_cell(
                df_central, df_dep, central_idx, dep_idx, col
            )
            comments.extend([[id_number] + get_soldier_info(dep_row) + [msg] for msg in comparison_comments])

    return df_central, df_dep, comments



def find_and_add_missing_rows(df_central, df_dep):
    comments = []

    # --- ניקוי וסידור מ.א. ---
    df_central["מ.א."] = clean_id_column(df_central["מ.א."])
    df_dep["מ.א."] = clean_id_column(df_dep["מ.א."])

    # --- בדיקת כפילויות מ.א. ---
    comments.extend(check_duplicate_ids(df_central, df_dep))

    # --- בדיקת תקינות מסגרות ראשיות ---
    comments.extend(check_valid_departments(df_central, "מרוכז"))
    comments.extend(check_valid_departments(df_dep, "פלוגתי"))

    # --- השוואת מזהים ---
    ids_central = set(df_central["מ.א."])
    ids_dep = set(df_dep["מ.א."])
    common_cols = [col for col in df_central.columns if col in df_dep.columns and col != "מ.א."]

    # --- הוספת חיילים חסרים למרוכז ---
    df_central, added_comments_central = add_missing_from_dep(df_central, df_dep, ids_central, ids_dep, common_cols)
    comments.extend(added_comments_central)

    # --- הוספת חיילים חסרים למחלקתי ---
    df_dep, added_comments_dep = add_missing_from_central(df_dep, df_central, ids_dep, ids_central, common_cols)
    comments.extend(added_comments_dep)

    # --- השוואת ערכים בסיסיים בין רשומות משותפות ---
    compare_comments = compare_shared_basic_fields(df_central, df_dep, ids_central & ids_dep)
    comments.extend(compare_comments)

    return df_central, df_dep, comments


def update_shabzak(df_central, df_dep, is_PALSAM=False):
    df_central = df_central.copy()
    df_dep = df_dep.copy()

    all_comments = []

    # שלב 1: הוספת שורות חסרות
    df_central, df_dep, comments_missing = find_and_add_missing_rows(df_central, df_dep)
    all_comments.extend(comments_missing)

    # שלב 2: השוואה ועדכון ערכים
    df_central, df_dep, comments_updates = compare_and_update_values(df_central, df_dep)
    all_comments.extend(comments_updates)

    # יצירת DataFrame להערות
    comments_df = pd.DataFrame(all_comments, columns=["מ.א.", "שם פרטי", "שם משפחה", "מסגרת ראשית", "מסגרת משנית", "הערה"])
    # הוספת רמת דחיפות וצבע
    comments_df["Priority"], comments_df["Color"] = zip(*comments_df["הערה"].map(get_priority_and_color))

    # מחזירה קובץ חדש ומעודכן מחולק לגליונות לפי המקורי
    if is_PALSAM:
        df_dep = split_to_sheets(df_dep)
        dep_xlsx = dict_to_excel_bytes(df_dep)
    else:
        dep_xlsx = to_excel_bytes(df_dep, "שבצק פלוגתי")

    return {
        "df_central": df_central,
        "df_dep": df_dep,
        "comments_df": comments_df,
        "central_xlsx": to_excel_bytes(df_central, "שבצק מרוכז"),
        "dep_xlsx": dep_xlsx,
        "comments_xlsx": to_colored_excel(comments_df, "הערות")
    }


################# ???????? #################

##################################### Platform #################################

# --- Settings ---
st.set_page_config(page_title="השליש האוטומטי", layout="centered")
st.markdown("""
    <style>
    body, .css-18e3th9, .css-1d391kg, .stTextInput, .stButton, .stSelectbox, .stFileUploader {
        direction: rtl;
        text-align: right;
    }
    .stMarkdown, .stText, .stSubheader, .stHeader {
        direction: rtl;
        text-align: right;
    }
    th, td {
        direction: rtl !important;
        text-align: right !important;
    }
    </style>
""", unsafe_allow_html=True)


if "page" not in st.session_state:
    st.session_state.page = "home"

# --- Home ---
if st.session_state.page == "home":
    st.markdown("<h1 style='text-align: center;'>השליש האוטומטי - מגן יהונתן 8552</h1>", unsafe_allow_html=True)

    center_col = st.columns([1, 1, 2, 1, 1])[2]
    with center_col:
        st.image("gdud8552.jpg", width=200)

    col_center = st.columns([1, 1, 2, 1, 1])[2]
    with col_center:
        if st.button("📆 עדכון שבצ\"ק יומי", use_container_width=True):
            go_to("daily_update")

# --- Daily update ---
elif st.session_state.page == "daily_update":
    st.markdown("<h1 style='text-align: center;'>עדכון שבצ\"ק יומי - מגן יהונתן 8552</h1>", unsafe_allow_html=True)


    center_col = st.columns([1, 1, 2, 1, 1])[2]
    with center_col:
        st.image("gdud8552.jpg", width=200)

    # --- main file ---
    central_file = st.file_uploader("בחר שבצק מרוכז", type=["xlsx", "xls", "csv"])

    # --- departments files ---
    dep_file = st.file_uploader("בחר שבצק פלוגתי", type=["xlsx", "xls", "csv"])

    # --- Run ---
    if st.button("🚀 עדכן שבצק"):
        if not central_file or not dep_file:
            st.error("יש להעלות גם קובץ מרוכז וגם לפחות קובץ מחלקתי אחד.")
        else:
            if "df_central" not in st.session_state:
                try:
                    if central_file.name.endswith("csv"):
                        df_central = pd.read_csv(central_file)
                    else:
                        df_central = pd.read_excel(central_file, sheet_name="שבצק מרוכז")
                except ValueError as e:
                    st.error("⚠️ לא נמצא גיליון בשם 'שבצק מרוכז'. ודא שהשם מדויק.")
                    st.stop()
                except Exception as e:
                    st.error(f"שגיאה בטעינת שבצק מרוכז: {e}")
                    st.stop()
            else:
                df_central = st.session_state["df_central"]
            is_PALSAM = any(keyword in dep_file.name for keyword in ["פלסם","פלס\"ם", "פלס_ם"])
            if is_PALSAM:
                st.success("📦 זוהה קובץ פלס\"ם - ההרצה תתבצע בהתאם.")
                df_dep = merge_all_sheets(dep_file)
            else:
                try:
                    if dep_file.name.endswith("csv"):
                        df_dep = pd.read_csv(dep_file)
                    else:
                        df_dep = pd.read_excel(dep_file)
                except Exception as e:
                    st.warning(f"שגיאה בטעינת הקובץ {dep_file.name}: {e}")

            st.success("✅ הקבצים נטענו בהצלחה. מוכן להריץ בדיקות.")

            valid_values = get_valid_values_by_filename(dep_file.name)

            result = update_shabzak(df_central, df_dep, is_PALSAM)

            st.session_state["download_ready"] = True
            st.session_state["central_xlsx"] = result["central_xlsx"]
            st.session_state["dep_xlsx"] = result["dep_xlsx"]
            st.session_state["comments_df"] = result["comments_df"]
            st.session_state["comments_xlsx"] = result["comments_xlsx"]
            st.session_state["df_central"] = result["df_central"]

            #  שמור הערות מצטברות
            if "all_comments_df" in st.session_state:
                st.session_state["all_comments_df"] = pd.concat(
                    [st.session_state["all_comments_df"], result["comments_df"]],
                    ignore_index=True
                )
            else:
                st.session_state["all_comments_df"] = result["comments_df"]

    if st.session_state.get("download_ready"):
        #st.subheader("📥 הורדת קבצים מעודכנים")
        st.markdown("<h3 style='text-align: right;'>📥 הורדת קבצים מעודכנים</h3>", unsafe_allow_html=True)

        col1, col2, col3 = st.columns(3)

        central_filename = Path(central_file.name).stem + "_מעודכן.xlsx"
        dep_filename = Path(dep_file.name).stem + "_מעודכן.xlsx"

        with col1:
            st.download_button(
                label="📁 הורד שבצ\"ק מרוכז",
                data=st.session_state["central_xlsx"],
                file_name=central_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col2:
            st.download_button(
                label="📁 הורד שבצ\"ק פלוגתי",
                data=st.session_state["dep_xlsx"],
                file_name=dep_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col3:
            st.download_button(
                label="📝 הורד דוח הערות",
                data=to_colored_excel(st.session_state["all_comments_df"], "הערות"),
                file_name="דוח_הערות.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        #st.subheader("📋 תצוגת דוח הערות")
        st.markdown("<h3 style='text-align: right;'>📋 תצוגת דוח הערות</h3>", unsafe_allow_html=True)

        #st.dataframe(st.session_state["comments_df"])
        render_comments_table(st.session_state["all_comments_df"])

        if st.button("עדכון פלוגה נוספת"):
            st.session_state.pop("download_ready", None)
            st.session_state.pop("central_xlsx", None)
            st.session_state.pop("dep_xlsx", None)
            st.session_state.pop("comments_xlsx", None)
            go_to("daily_update")

    if st.button("⬅️ חזרה למסך הראשי"):
        st.session_state.pop("download_ready", None)
        st.session_state.pop("central_xlsx", None)
        st.session_state.pop("dep_xlsx", None)
        st.session_state.pop("comments_xlsx", None)
        st.session_state.pop("comments_df", None)
        st.session_state.pop("df_central",None)

        go_to("home")

